import os
import re
from openpyxl import Workbook, load_workbook

def rename_folders_to_asin(base_folder='.'):
    asin_list = []
    log = []
    asin_pattern = re.compile(r'^([A-Z0-9]{10})([\s_\-\(]).*')
    # Duyệt toàn bộ thư mục, nhận diện cả folder đã đúng mã ASIN
    for name in os.listdir(base_folder):
        full_path = os.path.join(base_folder, name)
        if os.path.isdir(full_path):
            # Nếu là ASIN chuẩn rồi thì chỉ lưu vào danh sách, không cần đổi tên
            if re.fullmatch(r'[A-Z0-9]{10}', name):
                asin_list.append(name)
            else:
                match = asin_pattern.match(name)
                if match:
                    asin = match.group(1)
                    new_path = os.path.join(base_folder, asin)
                    if not os.path.exists(new_path):
                        os.rename(full_path, new_path)
                        log.append(f"🔁 Đổi tên: '{name}' → '{asin}'")
                    else:
                        log.append(f"⚠️ Thư mục '{asin}' đã tồn tại, bỏ qua '{name}'")
                    asin_list.append(asin)
                else:
                    log.append(f"⏩ Bỏ qua không khớp ASIN: {name}")
    return asin_list, log

def save_asins_to_excel(asin_list, excel_file='all.xlsx'):
    log = []
    if os.path.exists(excel_file):
        wb = load_workbook(excel_file)
        ws = wb.active
    else:
        wb = Workbook()
        ws = wb.active
        ws.append(["ASIN"])
    for asin in asin_list:
        ws.append([asin])
    wb.save(excel_file)
    log.append(f"📄 Đã lưu {len(asin_list)} ASIN vào '{excel_file}'")
    return log

def rename_main_web(current_folder='.', excel_file='all.xlsx'):
    asin_list, log1 = rename_folders_to_asin(current_folder)
    # Nếu không có ASIN do rename, vẫn kiểm tra lại các folder hợp lệ
    if not asin_list:
        # Quét lại tất cả thư mục hợp lệ (10 ký tự)
        asin_list = [name for name in os.listdir(current_folder)
                     if os.path.isdir(os.path.join(current_folder, name)) and re.fullmatch(r'[A-Z0-9]{10}', name)]
    log2 = []
    if asin_list:
        log2 = save_asins_to_excel(asin_list, excel_file)
    else:
        log2.append("❗ Không có thư mục hợp lệ để đổi tên và ghi Excel.")
    return log1 + log2
